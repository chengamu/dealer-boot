from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from agents.fabric_cutting.models import BusinessRules, CuttingRequest, Material, Piece, WrapMode


_SIZE_PATTERN = re.compile(
    r"(?P<width>\d+(?:\.\d+)?)\s*(?:m|米)?\s*[*xX×]\s*(?P<height>\d+(?:\.\d+)?)\s*(?:m|米)?"
)
_MATERIAL_PATTERN = re.compile(r"(?P<width>\d+(?:\.\d+)?)\s*(?:m|米)\s*(?:门幅|宽度|原材料|面料)?")
_ROLL_LENGTH_PATTERN = re.compile(
    r"(?:可用长度|面料可用长度|原材料长度|每卷长度|卷长|总长|长度)?\s*[:：]?\s*"
    r"(?P<total>\d+(?:\.\d+)?)\s*-\s*"
    r"(?P<head>\d+(?:\.\d+)?)\s*-\s*"
    r"(?P<tail>\d+(?:\.\d+)?)"
)


def parse_text_fallback(text: str) -> dict[str, Any]:
    materials: list[Material] = []
    pieces: list[Piece] = _parse_tabular_pieces(text)
    for match in _MATERIAL_PATTERN.finditer(text):
        width = float(match.group("width"))
        if width not in [material.width for material in materials]:
            materials.append(Material(width=width, id=f"{width:g}m"))

    start_index = len(pieces) + 1
    for index, match in enumerate(_SIZE_PATTERN.finditer(text), start=start_index):
        width = float(match.group("width"))
        height = float(match.group("height"))
        quantity = _parse_quantity_near(text[match.end() : match.end() + 16])
        pieces.append(Piece(id=str(index), width=width, height=height, quantity=quantity))

    wrap_mode = _parse_wrap_mode(text)
    roll_lengths = _parse_roll_lengths(text)
    return CuttingRequest(
        materials=materials or [Material(id="2m", width=2), Material(id="3m", width=3)],
        pieces=pieces,
        business_rules=BusinessRules(wrap_mode=wrap_mode, **roll_lengths),
    ).model_dump()


def _parse_tabular_pieces(text: str) -> list[Piece]:
    pieces: list[Piece] = []
    for line in text.splitlines():
        compact = line.strip()
        if not compact or any(header in compact.lower() for header in ["宽度", "高度", "width", "height"]):
            continue
        if any(keyword in compact for keyword in ["门幅", "可用长度", "卷长", "总长", "每卷长度", "原材料长度"]):
            continue
        numbers = re.findall(r"\d+(?:\.\d+)?", compact)
        if len(numbers) < 2:
            continue
        width = float(numbers[0])
        height = float(numbers[1])
        quantity = int(float(numbers[2])) if len(numbers) >= 3 else 1
        pieces.append(Piece(id=str(len(pieces) + 1), width=width, height=height, quantity=quantity))
    return pieces


def parse_excel_bytes(content: bytes) -> CuttingRequest:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 内容为空")

    header_row_index, headers = _find_header_row(rows)
    if header_row_index is None:
        raise ValueError("未找到包含 宽度/高度/数量 的表头")

    width_col = _find_col(headers, ["宽度", "宽", "width"])
    height_col = _find_col(headers, ["高度", "高", "height"])
    quantity_col = _find_col(headers, ["数量", "qty", "quantity"])
    id_col = _find_col(headers, ["编号", "序号", "id"])

    materials = _extract_materials(rows)
    pieces: list[Piece] = []
    for row_index, row in enumerate(rows[header_row_index + 1 :], start=1):
        width = _to_float(_cell(row, width_col))
        height = _to_float(_cell(row, height_col))
        if width is None or height is None:
            continue
        quantity = int(_to_float(_cell(row, quantity_col)) or 1) if quantity_col is not None else 1
        piece_id = str(_cell(row, id_col) or row_index) if id_col is not None else str(row_index)
        pieces.append(Piece(id=piece_id, width=width, height=height, quantity=quantity))

    return CuttingRequest(
        materials=materials or [Material(id="2m", width=2), Material(id="3m", width=3)],
        pieces=pieces,
        business_rules=BusinessRules(),
    )


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int | None, list[str]]:
    for index, row in enumerate(rows[:20]):
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        has_width = any(value in header for header in headers for value in ["宽度", "width"])
        has_height = any(value in header for header in headers for value in ["高度", "height"])
        if has_width and has_height:
            return index, headers
    return None, []


def _find_col(headers: list[str], names: list[str]) -> int | None:
    for index, header in enumerate(headers):
        if any(name in header for name in names):
            return index
    return None


def _extract_materials(rows: list[tuple[Any, ...]]) -> list[Material]:
    materials: list[Material] = []
    seen: set[float] = set()
    for row in rows[:30]:
        text = " ".join(str(cell) for cell in row if cell is not None)
        for match in _MATERIAL_PATTERN.finditer(text):
            width = float(match.group("width"))
            if width not in seen:
                seen.add(width)
                materials.append(Material(id=f"{width:g}m", width=width))
    return materials


def _parse_quantity_near(text: str) -> int:
    match = re.search(r"(?:数量|qty)?\s*[:：]?\s*(\d+)", text)
    return int(match.group(1)) if match else 1


def _parse_wrap_mode(text: str) -> WrapMode | None:
    compact = re.sub(r"\s+", "", text)
    if (
        "都不包布" in compact
        or "不包布" in compact and "下轨包布" not in compact and "罩壳包布" not in compact
        or "不包" in compact and "下轨包" not in compact and "罩壳包" not in compact and "都包" not in compact
        or compact in {"不包", "不包了"}
        or "包布方式不包" in compact
    ):
        return WrapMode.NO_COVER
    if "都包布" in compact or "罩壳+下轨" in compact:
        return WrapMode.BOTH_COVER
    if "罩壳包布" in compact:
        return WrapMode.HEADBOX_COVER
    if "下轨包布" in compact:
        return WrapMode.BOTTOM_RAIL_COVER
    return None


def _parse_roll_lengths(text: str) -> dict[str, float]:
    match = _ROLL_LENGTH_PATTERN.search(text)
    if not match:
        return {}
    return {
        "roll_total_length": float(match.group("total")),
        "roll_head_waste": float(match.group("head")),
        "roll_tail_waste": float(match.group("tail")),
    }


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, int | float):
        return float(value)
    match = re.search(r"\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else None
